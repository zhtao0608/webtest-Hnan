# -*- coding:utf-8 -*-
import xlrd,xlwt,os
from Base import ReadConfig
from Base.Mylog import LogManager
from Common.function import convertDicList
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import time
import openpyxl
from Data.DataMgnt.DataOper import DataOper as Dto

Lable = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
         'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
         'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ',
         'CA','CB','CC','CD','CE','CF','CG','CH','CI','CJ','CK','CL','CM','CN','CO','CP','CQ','CR','CS','CT','CU','CV','CW','CX','CY','CZ'
         ]
workbook = None

logger = LogManager('OperExcel').get_logger_and_add_handlers(1,is_add_stream_handler=True, log_path=ReadConfig.log_path, log_filename=time.strftime("%Y-%m-%d")+'.log' )
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
rc = ReadConfig.ReadConfig("ngboss_config.ini")

class ExlOp(object):
    '''Openxls重新封装'''
    def __init__(self, file,sheet):
        self.file = file
        self.wb = load_workbook(self.file)
        self.ws = self.wb[sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列的所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 设置某个单元格的值
    def set_cell_value(self, row, colunm, cellvalue):
        try:
            self.ws.cell(row=row, column=colunm).value = cellvalue
            self.wb.save(self.file)
        except:
            self.ws.cell(row=row, column=colunm).value = "writefail"
            self.wb.save(self.file)


def create_workbook(fileName,sheet='测试结果',value = [["子台帐数据"],["tf_b_trade"],["tf_b_trade_2020"]]):
    '''
    创建工作表
    :param fileName: 只是文件名，不是完整路径
    :return:
    '''
    dataFilePath = ReadConfig.get_data_path() + fileName + '_%s.xlsx' % time.strftime("%Y%m%d%H%M%S")
    if not os.path.exists(dataFilePath):
        writeExcel(path=dataFilePath,value=value,sheet=sheet)
        return dataFilePath
    else:
        return dataFilePath

def open_excel(file):
    """
    打开Excel
    :param path: Excel路径
    :return:
    """
    global workbook
    if workbook == None:
        workbook = xlrd.open_workbook(file)
    return workbook

def get_sheet(sheetName):
    global workbook
    return workbook.sheet_by_name(sheetName)

def get_sheet_by_index(index):
    global workbook
    return workbook.sheet_by_index(index)

def get_row(sheet):
    return sheet.nrows

def get_col(sheet):
    return sheet.ncols

def get_cell_content(sheet, row, col):
    """
    获取单元格内容
    :param sheet: sheet名
    :param row: 行
    :param col: 列
    :return:
    """
    return sheet.cell_value(row, col)

def getColumnIndex(file,columnName,sheet='Sheet1'):
    xl = xlrd.open_workbook(file)
    # sheet = xl.sheet_by_index(index)
    sheet = xl.sheet_by_name(sheet)
    table = sheet.row_values(0)
    for i in range(len(table)):
        if table[i] == columnName:
            columnIndex = i
            break
    return columnIndex

def getRowIndex(file,value,sheet='Sheet1'):
    '''根据xls表中的内容获取所在的行'''
    xl = xlrd.open_workbook(file)
    # sheet = xl.sheet_by_index(index)
    sheet = xl.sheet_by_name(sheet)
    nrows = sheet.nrows
    clos = sheet.ncols
    print(nrows,clos)
    for i in range(1,nrows):
        for j in range(0,clos):
            if sheet.cell_value(i,j) == value:
                rowIndex = i
    return rowIndex

def get_xls(xls_name, sheet_name):
    """
    获取excel表中指定sheet数据，保存到列表中返回
    :param xls_name: excel文件名
    :param sheet_name: sheet表名
    :return:
    """
    cls = []
    xls_path = os.path.join(ReadConfig.data_path, xls_name)
    # print(xls_path)
    file = open_excel(xls_path)
    sheet = file.sheet_by_name(sheet_name)
    sheet_nrows = sheet.nrows
    for i in range(sheet_nrows):
        if sheet.row_values(i)[0] != u'case_name':
            cls.append(sheet.row_values(i))
    return cls

def get_datas(file,name):
    xl = xlrd.open_workbook(file)
    sheet = xl.sheet_by_name(name)
    items = sheet.row_values(0)
    datas = []
    for nrow in range(1,sheet.nrows):
        data = dict()
        values = sheet.row_values(nrow)
        for ncol in range(0,len(items)):
            data[items[ncol]]=values[ncol]
            datas.append(data)
        return datas

#     获取xls数据，并转换成字典List，格式处理成json
#     :param filename: 完整路径
#     :param index: xls表格sheet页，如第一个sheet ,则index = 0
#     :return:数组，每个元素为dic
#     [{'leave': '昆明', 'arrive': '长沙', 'leave_date': '2020-04-30'},
#     {'leave': '上海', 'arrive': '南京', 'leave_date': '2020-05-30'},
#     {'leave': '北京', 'arrive': '武汉', 'leave_date': '2020-04-22'}]

def get_exceldata(filename,index=0):
	boot_data = xlrd.open_workbook(filename)
	book_sheet = boot_data.sheet_by_index(index)
	rowsnum = book_sheet.nrows
	# print("列表行数："+str(rowsnum))
	#第一行当做字段的键
	rows0 = book_sheet.row_values(0)
	rows0_num = len(rows0)
	# print("列表列数:"+str(rows0_num))
	list = []
    #获取列表数据，并转换成字典List，格式处理成json
	for i in range(1,rowsnum):
		rows_data = book_sheet.row_values(i)  ###取第一行的值作为列表
		rows_dic = {}
		for y in range(0,rows0_num):
			rows_dic[rows0[y]]=rows_data[y]
		list.append(rows_dic)
	return list

def read_excel_byList(file_name, index=0):
    '''按行读取xls，并将xls内容（包含表头），返回list
    ['leave', 'arrive', 'leave_date', '昆明', '长沙', '2020-04-30', '上海', '南京', '2020-05-30', '北京', '武汉', '2020-04-22']
    '''
    xls = xlrd.open_workbook(file_name)
    sheet = xls.sheet_by_index(index)
    print(sheet.nrows)
    print(sheet.ncols)
    data = []
    for j in range(sheet.nrows):
        for i in range(sheet.ncols):
            data.append(sheet.row_values(j)[i])
            print(sheet.row_values(j)[i])
            # print("函数体类")
        # dic[j] = data
    print(data)
    return data

def read_excel(filename,index=0):
    ''' 按xls列逐个读取，按列返回字典，列数为key ,每列内容为value
    :param filename:
    :param index:
    :return:
    {0: ['leave', '昆明', '上海', '北京'],
    1: ['arrive', '长沙', '南京', '武汉'],
    2: ['leave_date', '2020-04-30', '2020-05-30', '2020-04-22']}
    '''
    xls = xlrd.open_workbook(filename)
    sheet = xls.sheet_by_index(index)
    print(sheet.nrows)
    print(sheet.ncols)
    dic = {}
    for j in range(sheet.ncols):
        data =[]
        for i in range(sheet.nrows):
            data.append(sheet.row_values(i)[j])
        dic[j]=data
    return dic

def read_xls_by_row(filename,index=0):
    '''按xls行读取，返回dic,每行内容做为一个数组
    {
    0: ['leave', 'arrive', 'leave_date'],
    1: ['昆明', '长沙', '2020-04-30'],
    2: ['上海', '南京', '2020-05-30'],
    3: ['北京', '武汉', '2020-04-22']}
    '''
    xls = xlrd.open_workbook(filename)
    sheet = xls.sheet_by_index(index)
    print(sheet.nrows)
    print(sheet.ncols)
    dic = {}
    for i in range(sheet.nrows):
        data = []
        for j in range(sheet.ncols):
            data.append(sheet.row_values(i)[j])
        dic[i] = data
    return dic

def writeExcel(path, value, sheet):
    '''
    :param sheet:sheet的名称
    :param path:文件的名字和路径
    :param value1: 写入的数据
    :return:
    '''
    book = openpyxl.Workbook()
    sheet1 = book.active
    sheet1.title = sheet
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet1.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    book.save(path)
    book.close()
    print("写入数据成功！")



def write_excel_xls(file, sheet_name, value):
    '''新建表格写入'''
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(file)  # 保存工作簿
    print("xls格式表格写入数据成功！")


def writeToExcel(fileName,sheetName,data):
    '''按数据列写入'''
    # file_name = 'test_1.xls'
    # if not os.path.exists(fileName):
    #     workbook = xlwt.Workbook(encoding='ascii')
    #     worksheet = workbook.add_sheet(sheetName,cell_overwrite_ok=True)
    #     workbook.save(fileName)
    rb = xlrd.open_workbook(fileName)
    sheets = rb.sheet_names()
    if sheetName in sheets:
        table = rb.sheet_by_name(sheetName)
        nrows = table.nrows
        ncols = table.ncols
        print('获取工作表{}存在，sheet名是:{},行数:{},列数:{}'.format(fileName,sheetName,nrows,ncols))
    wb = copy(rb)
    # 新的sheet页面
    if not sheetName in sheets:
        worksheet = wb.add_sheet(sheetName,cell_overwrite_ok=True)
        for i in range(0,len(data)):
            for j in range(0, len(data[i])):
                worksheet.write(j, i, data[i][j])  # 像表格中写入数据（对应的行和列)
                # worksheet.write(i, j, data[i][j])  # 像表格中写入数据（对应的行和列)
        wb.save(fileName)
    else:   #如果sheet已存在，则直接追加写入
        ws = wb.get_sheet(sheetName)
        newData = data[1:] #去掉字段列表()
        # k = nrows    #补充写入数据的初始行
        k = ncols    #补充写入数据的初始行
        for i in range(0,len(newData)):
            for j in range(0, len(newData[i])):
                ws.write(j, k, newData[i][j])  # 像表格中写入数据（对应的行和列)
                # ws.write(k, j, newData[i][j])  # 像表格中写入数据（对应的行和列)
            k = k+1
        wb.save(fileName)


def writeToExcelByRows(fileName,sheetName,data):
    '''按行写入'''
    rb = xlrd.open_workbook(fileName)
    sheets = rb.sheet_names()
    if sheetName in sheets:
        table = rb.sheet_by_name(sheetName)
        nrows = table.nrows
        ncols = table.ncols
        print('获取工作表{}存在，sheet名是:{},行数:{},列数:{}'.format(fileName,sheetName,nrows,ncols))
    wb = copy(rb)
    # 新的sheet页面
    if not sheetName in sheets:
        worksheet = wb.add_sheet(sheetName,cell_overwrite_ok=True)
        for i in range(0,len(data)):
            for j in range(0, len(data[i])):
                worksheet.write(i, j, data[i][j])  # 像表格中写入数据（对应的行和列)
        wb.save(fileName)
    else:   #如果sheet已存在，则直接追加写入
        ws = wb.get_sheet(sheetName)
        newData = data[1:] #去掉字段列表()
        k = nrows    #补充写入数据的初始行
        # k = ncols    #补充写入数据的初始行
        for i in range(0,len(newData)):
            for j in range(0, len(newData[i])):
                # ws.write(j, k, newData[i][j])  # 像表格中写入数据（对应的行和列)
                ws.write(k, j, newData[i][j])  # 像表格中写入数据（对应的行和列)
            k = k+1
        wb.save(fileName)


def writeDictToXls(fileName,sheetName,data):
    '''
    data 是合并后的字典类型
    :param fileName:
    :param sheetName: sheet页名称
    :param data: 这里的数据类型是经过dict合并过后的数据结构
    :return:
    '''
    if not os.path.exists(fileName):
        workbook = xlwt.Workbook(encoding='ascii')
        worksheet = workbook.add_sheet('测试结果',cell_overwrite_ok=True)
        workbook.save(fileName)
    rb = xlrd.open_workbook(fileName)
    sheets = rb.sheet_names()
    wb = copy(rb)
    keys = []
    values = []
    ##先处理下字典
    for key, value in data.items():  # 先首列写入key，
        keys.append(key)
        values.append(value)
    print(keys, values)
    # 新的sheet页面
    print('####获取所有sheet',sheets)
    if not sheetName in sheets:
        ws = wb.add_sheet(sheetName,cell_overwrite_ok=True)
    else:
        ws = wb.get_sheet(sheetName)
    print('开始写入数据:',values)
    for i in range(0,len(keys)):
        ws.write(i,0, keys[i])  # 先首列写入key
    ##注意要把数据格式处理下
    for j in range(0,len(values)):
        if isinstance(values[j],list):  #列表里的所有元素转换成str后写入
            for v in range(0,len(values[j])):
                data = values[j][v]
                if isinstance(data,datetime.datetime):
                    data = data.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(data,int):
                    data = str(values[j][v])
                ws.write(j,v+1, data)  # 先第2列开始写入value
        else:       #返回不是list就是单个值
            if isinstance(values[j],datetime.datetime):
                values[j] = values[j].strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(values[j],int):
                values[j] = str(values[j])
            ws.write(j, j+1, values[j])  # 先第2列开始写入value
    wb.save(fileName)
    print('数据写入xls完毕!xls文件名:{}'.format(fileName))


def write_dict_xls(inputData,sheetName,outPutFile):
    '''
    :param inputData: 列表，含有多个字典;例如：[{'key_a':'123'},{'key_b':'456'}]
    :param outPutFile: 输出文件名，例如：'data.xlsx'
    :return:
    '''
    wb = Workbook()
    # wb = load_workbook(outPutFile)
    wb.create_sheet(sheetName)
    print('########',outPutFile)
    sheet = wb.active
    sheet.title = sheetName
    item_0 = inputData[0]
    print('要写入xls的数据：',inputData)
    i = 0
    for key in item_0.keys():
        sheet[Lable[i]+str(1)].value = key
        i = i+1
    j = 1
    for item in inputData:
        k = 0
        for key in item:
            if isinstance(item[key],datetime.datetime): ##如果是时间格式要转换成字符
                item[key] = item[key].strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(item[key],int):
                item[key] = str(item[key])
            sheet[Lable[k]+str(j+1)].value = item[key]
            k = k+1
        j = j+1
    wb.save(outPutFile)
    print('数据写入xls完毕!xls文件名:{}'.format(outPutFile))



def write_excel_append(file,row,col,value,index=0):
    '''追加写入xls，在xls指定行列写入数据'''
    workbook = xlrd.open_workbook(file,formatting_info=True)  # 打开工作簿
    worksheet = workbook.sheet_by_index(index)  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    # 将xlrd对象拷贝转化为xlwt对象
    new_workbook = copy(workbook)
    new_worksheet = new_workbook.get_sheet(index)  # 获取转化后工作簿中的第一个表格
    new_worksheet.write(row, col, value) # 像指定行列写入数据
    new_workbook.save(file)  # 保存工作簿
    print("xls格式表格【追加】写入数据成功！")


def write_xlsBycolName_append(file,row,colName,value,index=0):
    '''追加写入xls，在xls指定行列写入数据'''
    filename = file
    '''
    打开xls时是否保留原来格式，这里兼容性存在问题，这里做个兼容处理
    （保留格式打开的时候出现异常就不保留格式了）
    '''
    try:
        workbook = xlrd.open_workbook(filename,formatting_info=True)  # 打开工作簿
    except:
        workbook = xlrd.open_workbook(filename)  # 打开工作簿
    # workbook = xlrd.open_workbook(filename)  # 打开工作簿
    worksheet = workbook.sheet_by_index(index)  # 获取工作簿中所有表格中的的第一个表格
    # rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    # 将xlrd对象拷贝转化为xlwt对象
    new_workbook = copy(workbook)
    new_worksheet = new_workbook.get_sheet(index)  # 获取转化后工作簿中的第一个表格
    colname = colName
    col = getColumnIndex(file,colname,index)  #根据colName获取对应的列数
    oldvalue = worksheet.cell_value(row,col)
    if not isinstance(oldvalue, str):
        oldvalue = str(oldvalue)
    if not isinstance(value, str):
        value = str(value)
    values = oldvalue + '\n' + value
    new_worksheet.write(row, col, values) # 像指定行列写入数据
    new_workbook.save(file)  # 保存工作簿
    print("xls表格【追加】写入数据成功！")

def write_xlsBycolName(file,row,colName,value,sheet='sheet1'):
    '''追加写入xls，在xls指定行列写入数据'''
    filename = file
    '''
    打开xls时是否保留原来格式，这里兼容性存在问题，这里做个兼容处理
    （保留格式打开的时候出现异常就不保留格式了）
    '''
    try:
        workbook = xlrd.open_workbook(filename,formatting_info=True)  # 打开工作簿
    except:
        workbook = xlrd.open_workbook(filename)  # 打开工作簿
    # worksheet = workbook.sheet_by_index(index)  # 获取工作簿中所有表格中的的第一个表格
    worksheet = workbook.sheet_by_name(sheet)  # 获取工作簿中所有表格中的的第一个表格
    new_workbook = copy(workbook)
    new_worksheet = new_workbook.get_sheet(sheet)  # 获取转化后工作簿中的第一个表格
    colname = colName
    col = getColumnIndex(file,colname,sheet)  #根据colName获取对应的列数
    # oldvalue = worksheet.cell_value(row,col)
    # if not isinstance(oldvalue, str):
    #     oldvalue = str(oldvalue)
    if not isinstance(value, str):
        value = str(value)
    values =  value
    new_worksheet.write(row, col, values) # 像指定行列写入数据
    new_workbook.save(file)  # 保存工作簿
    print("xls表格【追加】写入数据成功！")


if __name__ == '__main__':
    CreatePersonUserMenu = Dto().getSysMenuByParentFuncId(parentFuncId='crm9100')
    logger.info('========个人开户菜单列表：{}'.format(CreatePersonUserMenu))
    dataFile = create_workbook(fileName='个人业务菜单冒烟测试', value=convertDicList(CreatePersonUserMenu), sheet='开户业务')
    print('#####x写入的xls文件名:', dataFile)

    EPostInfoMenu = Dto().getSysMenuByParentFuncId(parentFuncId='crm9A00')
    logger.info('========个人电子发票菜单列表：{}'.format(EPostInfoMenu))
    writeToExcelByRows(fileName=dataFile, sheetName='电子发票', data=convertDicList(EPostInfoMenu))
    print('#####写入的xls文件名:', dataFile)

    ArticMenu = Dto().getSysMenuByParentFuncId(parentFuncId='crm9830')
    logger.info('========引商入柜业务菜单列表：{}'.format(ArticMenu))
    writeToExcelByRows(fileName=dataFile, sheetName='引商入柜业务', data=convertDicList(ArticMenu))
    print('#####写入的xls文件名:', dataFile)

    PersonDailyMenu = Dto().getSysMenuByParentFuncId(parentFuncId='crm9200')
    logger.info('========引商入柜业务菜单列表：{}'.format(PersonDailyMenu))
    writeToExcelByRows(fileName=dataFile, sheetName='日常业务', data=convertDicList(ArticMenu))
    print('#####写入的xls文件名:', dataFile)
    write_xlsBycolName(file=dataFile, row=getRowIndex(file=dataFile, value='crm9113',sheet='开户业务'), colName='TEST_RESULT',
                       value='Pass',sheet='开户业务')
    write_xlsBycolName(file=dataFile, row=getRowIndex(file=dataFile, value='crm9835',sheet='日常业务'), colName='TEST_RESULT',
                       value='Fail',sheet='日常业务')
    # file = ReadConfig.data_path + 'testDatas_{}.xlsx'.format(time.strftime("%Y%m%d%H%M%S"))
    # datas = {'TRADE_ID': [3120082587858316, 3120082587858316], 'ACCEPT_MONTH': [8, 8], 'USER_ID': [3120082500014516, 3120082500014516], 'USER_ID_A': [-1, -1], 'PACKAGE_ID': [32953733, 99966954], 'PRODUCT_ID': [32811359, 32811359], 'OFFER_TYPE': ['D', 'D'], 'OFFER_ID': [130032532282, 130099665664], 'DISCNT_CODE': [32532282, 99665664], 'SPEC_TAG': ['0', '0'], 'RELATION_TYPE_CODE': [None, None], 'INST_ID': [3120082500029184, 3120082500029185], 'CAMPN_ID': [None, None], 'OLD_PRODUCT_ID': [None, None], 'OLD_PACKAGE_ID': [None, None], 'START_DATE': [datetime.datetime(2020, 8, 25, 20, 0, 4), datetime.datetime(2020, 8, 25, 20, 0, 4)], 'END_DATE': [datetime.datetime(2050, 12, 31, 0, 0), datetime.datetime(2022, 7, 31, 23, 59, 59)], 'MODIFY_TAG': ['0', '0'], 'UPDATE_TIME': [datetime.datetime(2020, 8, 25, 20, 0, 4), datetime.datetime(2020, 8, 25, 20, 0, 4)], 'UPDATE_STAFF_ID': ['ITFTA114', 'ITFTA114'], 'UPDATE_DEPART_ID': ['17EFF', '17EFF'], 'OPER_CODE': [None, None], 'IS_NEED_PF': [None, None], 'CREATE_DATE': [datetime.datetime(2020, 8, 25, 20, 0, 4), datetime.datetime(2020, 8, 25, 20, 0, 4)], 'CREATE_STAFF_ID': ['ITFTA114', 'ITFTA114'], 'CREATE_DEPART_ID': ['17EFF', '17EFF'], 'DONE_CODE': [3120082587858316, 3120082587858316], 'REMARK': [None, None], 'RSRV_DATE1': [None, None], 'RSRV_DATE2': [None, None], 'RSRV_DATE3': [None, None], 'RSRV_NUM1': [None, None], 'RSRV_NUM2': [None, None], 'RSRV_NUM3': [None, None], 'RSRV_NUM4': [None, None], 'RSRV_NUM5': [None, None], 'RSRV_STR1': [None, None], 'RSRV_STR2': [None, None], 'RSRV_STR3': [None, None], 'RSRV_STR4': [None, None], 'RSRV_STR5': [None, None], 'RSRV_TAG1': [None, None], 'RSRV_TAG2': [None, None], 'RSRV_TAG3': [None, None]}
    # writeDictToXls(fileName=file,sheetName='test',data=datas)
    # wb = open_excel(ReadConfig.data_path+"TestCase.xls")
    # ws = get_sheet_by_index(0)
    # row = get_row(ws)
    # col = get_col(ws)
    # cell_content = get_cell_content(ws, 3, 2)
    # print(type(cell_content))
    # print(ws.name, row, col, cell_content)
    # datas = get_exceldata(ReadConfig.data_path+"testdata2.xls",0)
    # offers_para = get_exceldata(ReadConfig.data_path+"UITest_GrpBusiOper.xls",0)[0]
    # print(type(offers_para))
    # print("业务受理参数：", offers_para)
    # for i in range(len(offers_para)):
    #     # print(offers_para[i])
    #     print(offers_para[i]['subofferList'])
    # index = int(offers_para.get('No'))  # 标识行号，后续写入xls使用
    # print("index:" + str(index))
    # groupId = str(offers_para.get('groupId'))
    # print("集团编码:" + groupId)
    # offerid = str(offers_para.get('mainoffer'))  # 集团主商品ID
    # subOfferList = offers_para.get('subofferList').replace(" ", "").split(',')
    # # subOfferList = dict_get(offers_para,'subofferList',None).split(',')
    # print("subOfferList:" ,subOfferList)
    # print(type(subOfferList))

    # xls = get_xls("TestCase.xls","intf_test")
    # print(xls)
    # value = ['ok']
    # print(len(value))
    # write_excel_append(ReadConfig.data_path+"testdata2.xls", value)
    # colIndex = getColumnIndex(ReadConfig.data_path + 'UITest_TestData.xls', 'PARAMS', index=0)
    # print('PARAS对应的列数：',colIndex)
    # file = ReadConfig.data_path+"UITest_data_20200525005.xls"
    # print(file)
    # write_xlsBycolName_append(file, 2, 'RESULT_INFO', 'testtest-nonono')  # 向xls模板指定行列写入结果

    # file = ReadConfig.data_path+"UITest_ShareActive.xls"
    # value = [{'NO': 1, 'ACCESS_NUM': '18708700011', 'SUBSCRIBER_INS_ID': '7012082036480364', 'FLOW_ID': None, 'RESULT_INFO': None}, {'NO': 2, 'ACCESS_NUM': '18708700018', 'SUBSCRIBER_INS_ID': '7015102653160225', 'FLOW_ID': None, 'RESULT_INFO': None}, {'NO': 3, 'ACCESS_NUM': '18708700048', 'SUBSCRIBER_INS_ID': '7014011344200264', 'FLOW_ID': None, 'RESULT_INFO': None}]
    # # write_dict_xls(inputData =value, sheetName='testData', outPutFile=file)
    # write_xlsBycolName_append(file, 1, '检查点', '测试测试')
    # write_xlsBycolName_append(file, 1, '交互流水号', '1231231231231233')

    # rowindex = getRowIndex(file=file,value='SubscriberOpenTest')
    # print('SubscriberOpenTest所在的行:',rowindex)
    # colIndex = getColumnIndex(file, 'PARAMS', index=0)
    # print('PARAS对应的列数：',colIndex)
    # write_xlsBycolName_append(file, rowindex, 'FLOWID', '2222222')













